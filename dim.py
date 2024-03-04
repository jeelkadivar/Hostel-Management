import qrcode
import openpyxl
from datetime import datetime

# Create a new Excel workbook or open an existing one
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.cell(row=1, column=1).value = "Content"
sheet.cell(row=1, column=2).value = "Date/Time"
row_num = 2

while True:
    # Scan the QR code and get its content
    img = qrcode.make('your-QR-code-here')
    content = str(qrcode.process(img)[0].data.decode())

    # Get the current date and time
    now = datetime.now()
    dt_string = now.strftime("%d/%m/%Y %H:%M:%S")

    # Write the content and date/time to the next available row in the Excel file
    sheet.cell(row=row_num, column=1).value = content
    sheet.cell(row=row_num, column=2).value = dt_string
    row_num += 1

    # Save the Excel file
    filename = "QR_codes.xlsx"
    workbook.save(filename)